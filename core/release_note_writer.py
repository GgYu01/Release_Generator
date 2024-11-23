# core/release_note_writer.py

import os
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from utils.logger import get_logger
from utils.exception_handler import AppException
from config.settings import config

logger = get_logger(__name__)


class ReleaseNoteWriter:
    """
    Class to generate and populate Release Note spreadsheets with commit and patch information.
    """

    def __init__(self, template_path: str, output_path: str):
        """
        Initialize the ReleaseNoteWriter.

        :param template_path: Path to the existing Release Note Excel file.
        :param output_path: Path where the updated Release Note Excel file will be saved.
        """
        self.template_path = template_path
        self.output_path = output_path
        self.workbook = None
        self.sheet = None

    def load_workbook(self):
        """
        Load the Release Note Excel workbook and select the active sheet.
        """
        try:
            self.workbook = load_workbook(self.template_path)
            self.sheet = self.workbook.active
            logger.info(f"Loaded Release Note template from {self.template_path}")
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            raise AppException(f"Error loading workbook: {e}")

    def save_workbook(self):
        """
        Save the workbook to the specified output path.
        """
        try:
            self.workbook.save(self.output_path)
            logger.info(f"Saved updated Release Note to {self.output_path}")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            raise AppException(f"Error saving workbook: {e}")

    def insert_commit_rows(self, num_rows: int):
        """
        Insert the specified number of rows starting from the second row.

        :param num_rows: Number of rows to insert.
        """
        self.sheet.insert_rows(2, num_rows)
        logger.debug(f"Inserted {num_rows} rows at row 2")

    def populate_release_notes(
        self,
        release_version: str,
        commit_infos: List[Dict],
        module_mappings: Dict[str, str],
        special_commits: List[Dict] = None
    ):
        """
        Populate the Release Note spreadsheet with commit and patch information.

        :param release_version: Name of the latest git TAG in ~/grt.
        :param commit_infos: List of commit information dictionaries.
        :param module_mappings: Mapping of repository names to module names.
        :param special_commits: List of special commit dictionaries, if any.
        """
        total_commits = len(commit_infos)
        logger.info(f"Total commits to write: {total_commits}")

        if total_commits == 0:
            logger.warning("No commits available to populate.")
            return

        self.insert_commit_rows(total_commits)

        # Start writing from row 2
        row_idx = 2

        for commit_info in commit_infos:
            repository_name = commit_info['repository']
            commit_message = commit_info['message']
            commit_id = commit_info['commit_id']
            patch_path = commit_info.get('patch_path', '')
            is_special = commit_info.get('is_special', False)
            module_name = module_mappings.get(repository_name, 'Unknown')

            # Handle the special commit mapping for grpower or nebula
            if is_special:
                if '] thyp-sdk: ' in commit_message or '] nebula-sdk: ' in commit_message:
                    module_name = 'grpower'
                elif '] tee: ' in commit_message:
                    module_name = 'nebula'
                # Update patch path for special commits
                patch_path = f"/patches/{module_name}/{commit_id}.patch"

            # Remove '/home/nebula' from the patch path for normal commits
            if patch_path.startswith('/home/nebula'):
                patch_path = patch_path.replace('/home/nebula', '')

            # Fill in the columns as per requirements
            self.sheet.cell(row=row_idx, column=1).value = release_version  # Column A
            self.sheet.cell(row=row_idx, column=2).value = commit_message   # Column B
            self.sheet.cell(row=row_idx, column=3).value = module_name      # Column C
            self.sheet.cell(row=row_idx, column=4).value = patch_path       # Column D
            self.sheet.cell(row=row_idx, column=5).value = ''               # Column E (Leave blank)
            self.sheet.cell(row=row_idx, column=6).value = commit_info.get('additional_info', '')  # Column F
            self.sheet.cell(row=row_idx, column=7).value = config.default_test_leader  # Column G

            # Columns H to N: Fill based on configuration or leave blank
            for col in range(8, 15):
                self.sheet.cell(row=row_idx, column=col).value = ''  # Adjust as needed

            self.sheet.cell(row=row_idx, column=15).value = commit_id  # Column O

            # Align text to left for all cells in the row
            for col in range(1, self.sheet.max_column + 1):
                cell = self.sheet.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal='left')

            logger.debug(f"Populated row {row_idx} with commit {commit_id}")
            row_idx += 1

        logger.info(f"Populated {total_commits} commits into the Release Note.")

    def generate_release_note(
        self,
        release_version: str,
        commit_infos: List[Dict],
        module_mappings: Dict[str, str],
        special_commits: List[Dict] = None
    ):
        """
        Main method to generate the Release Note.

        :param release_version: Name of the latest git TAG in ~/grt.
        :param commit_infos: List of commit information dictionaries.
        :param module_mappings: Mapping of repository names to module names.
        :param special_commits: List of special commit dictionaries, if any.
        """
        self.load_workbook()
        self.populate_release_notes(
            release_version,
            commit_infos,
            module_mappings,
            special_commits
        )
        self.save_workbook()