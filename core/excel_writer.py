# core/excel_writer.py

from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from config.settings import settings, RepositoryInfo, CommitInfo
from core.git_handler import GitHandler
from utils.logger import get_logger
from rich.console import Console
import datetime
import os
import re

# Regular expression to match illegal characters
ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\x00-\x08\x0B-\x0C\x0E-\x1F]'
)

class ExcelWriter:
    def __init__(self, output_path: str) -> None:
        self.output_path = Path(output_path)
        self.logger = get_logger('ExcelWriter')
        self.console = Console()
        self.workbook: Workbook
        self.worksheet: Worksheet
        self._initialize_workbook()

    def _initialize_workbook(self) -> None:
        if self.output_path.exists():
            self.workbook = load_workbook(self.output_path)
            self.logger.info(f"Loaded existing workbook from {self.output_path}")
            self.console.log(f"[green]Loaded existing workbook from {self.output_path}[/green]")
        else:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self._create_header()
            self.logger.info("Created new workbook")
            self.console.log("[green]Created new workbook[/green]")

    def _create_header(self) -> None:
        headers = [
            'Latest Git TAG',  # Column A
            'Commit Message',  # Column B
            'Parent Repository',  # Column C
            'Patch File',  # Column D
            'Topic Content',  # Column E
            'Specific Repos Last Commit IDs',  # Column F
            'Responsible Persons',  # Column G
            'Submission Time',  # Column H
            'Needs Porting',  # Column I
            'Porting Done',  # Column J
            'Send to Customer',  # Column K
            '',  # Column L (blank)
            '',  # Column M (blank)
            '',  # Column N (blank)
            'Commit ID'  # Column O
        ]
        self.worksheet.append(headers)
        self.logger.debug("Header row created")

    def write_commits(self, repositories: List[RepositoryInfo]) -> None:
        self.worksheet = self.workbook.active
        starting_row = self.worksheet.max_row + 1
        for repo in repositories:
            for commit in repo.commits:
                try:
                    row_data = self._prepare_row_data(repo, commit)
                    sanitized_row_data = [self._sanitize_string(cell) if isinstance(cell, str) else cell for cell in row_data]
                    self.worksheet.append(sanitized_row_data)
                    self.logger.debug(f"Written commit {commit.commit_id} to Excel")
                    self.console.log(f"[cyan]Written commit {commit.commit_id} to Excel[/cyan]")
                except Exception as e:
                    error_message = f"Error writing commit {commit.commit_id} to Excel: {e}"
                    self.logger.error(error_message)
                    self.console.log(f"[red]{error_message}[/red]")
                    continue  # Skip this commit and continue with the next one
        try:
            self.workbook.save(self.output_path)
            self.logger.info(f"Workbook saved to {self.output_path}")
            self.console.log(f"[green]Workbook saved to {self.output_path}[/green]")
        except Exception as e:
            error_message = f"Error saving workbook: {e}"
            self.logger.error(error_message)
            self.console.log(f"[red]{error_message}[/red]")

    def _prepare_row_data(self, repo: RepositoryInfo, commit: CommitInfo) -> List[Any]:
        # Column A: Latest Git TAG from /home/nebula/grt
        latest_git_tag = self._get_grt_latest_tag()
        # Column B: Commit Message
        commit_message = commit.message
        # Column C: Parent Repository Name (based on commit's parent_repos)
        parent_repo_names = '\n'.join(commit.parent_repos) if commit.parent_repos else self._get_parent_repo_name(repo)
        # Column D: Patch File Path
        patch_file = commit.patch_file or ''
        # Column E: Topic Related Content (configurable)
        topic_content = self._get_topic_content(commit)
        # Column F: Last Commit IDs from specific repositories
        last_commit_ids = self._get_specific_repo_last_commits()
        # Column G: Responsible Persons (configurable)
        responsible_persons = settings.responsible_person_info
        # Column H: Submission Time (configurable)
        submission_time = self._get_submission_time(commit)
        # Column I: Needs Porting (configurable)
        needs_porting = settings.porting_status_options.get('needs_porting', '')
        # Column J: Porting Done (configurable)
        porting_done = settings.porting_status_options.get('porting_done', '')
        # Column K: Send to Customer (configurable)
        send_to_customer = settings.porting_status_options.get('send_to_customer', '')
        # Column O: Commit ID
        commit_id = commit.commit_id

        row_data = [
            latest_git_tag,         # Column A
            commit_message,         # Column B
            parent_repo_names,      # Column C
            patch_file,             # Column D
            topic_content,          # Column E
            last_commit_ids,        # Column F
            responsible_persons,    # Column G
            submission_time,        # Column H
            needs_porting,          # Column I
            porting_done,           # Column J
            send_to_customer,       # Column K
            '',                     # Column L
            '',                     # Column M
            '',                     # Column N
            commit_id               # Column O
        ]

        return row_data

    def _sanitize_string(self, value: str) -> str:
        """Remove illegal characters from the string."""
        sanitized_value = ILLEGAL_CHARACTERS_RE.sub('', value)
        return sanitized_value

    def _get_grt_latest_tag(self) -> str:
        # Assuming grt repository is configured
        grt_repo = next((repo for repo in settings.repositories if repo.name == 'grt'), None)
        if grt_repo:
            git_handler = GitHandler(grt_repo.path)
            latest_tag, _ = git_handler.get_last_two_tags()
            self.logger.debug(f"Latest GRT tag: {latest_tag}")
            return latest_tag
        return ''

    def _get_parent_repo_name(self, repo: RepositoryInfo) -> str:
        if repo.parent:
            mapped_name = settings.parent_repo_mapping.get(repo.name, repo.parent)
            self.logger.debug(f"Parent repository for {repo.name}: {mapped_name}")
            return mapped_name
        else:
            self.logger.debug(f"No parent repository for {repo.name}, using repository name")
            return repo.name

    def _get_topic_content(self, commit: CommitInfo) -> str:
        # Placeholder for configurable topic content logic
        # User can customize this method
        topic_content = "Custom Topic Content"
        self.logger.debug(f"Topic content for commit {commit.commit_id}: {topic_content}")
        return topic_content

    def _get_specific_repo_last_commits(self) -> str:
        repos_to_check = [
            '/home/nebula/grpower/workspace/nebula/zircon',
            '/home/nebula/grpower/workspace/nebula/garnet'
        ]
        last_commits = []
        for repo_path in repos_to_check:
            try:
                git_handler = GitHandler(repo_path)
                last_commit_id = git_handler.get_latest_commit_id()
                relative_path = os.path.relpath(repo_path, '/home/nebula')
                last_commits.append(f"{relative_path}: {last_commit_id}")
                self.logger.debug(f"Last commit for {relative_path}: {last_commit_id}")
            except Exception as e:
                error_message = f"Error retrieving latest commit for {repo_path}: {e}"
                self.logger.error(error_message)
                self.console.log(f"[red]{error_message}[/red]")
                continue  # Skip this repository and continue with the next one
        return '\n'.join(last_commits)

    def _get_submission_time(self, commit: CommitInfo) -> str:
        time_format = settings.submission_time_format
        # Placeholder for actual commit time extraction
        submission_time = datetime.datetime.now().strftime(time_format)
        self.logger.debug(f"Submission time for commit {commit.commit_id}: {submission_time}")
        return submission_time
